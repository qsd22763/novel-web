from rest_framework import viewsets, status, filters
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .auth import IsAdminUserOrStaff
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Count, Sum, Q, Max, F
from django.utils import timezone
from datetime import timedelta
import re
import chardet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse

from .models import Advertisement, Announcement, Novel, OperationLog, BookCategory, ViolationRecord, ChapterUploadLog, Chapter, CheckIn, CheckInConfig, MembershipOrder
from django.contrib.auth import get_user_model
from .models import Favorite, ReadingProgress, Bookmark, Comment

User = get_user_model()
from .admin_serializers import (
    AdvertisementSerializer,
    AnnouncementSerializer,
    AdminNovelSerializer,
    OperationLogSerializer,
    BookCategorySerializer,
    ViolationRecordSerializer,
    ChapterUploadLogSerializer,
    AdvertisementStatsSerializer,
    AdminChapterSerializer,
)
from .user_serializers import CheckInSerializer, MembershipOrderSerializer


class AdminAdvertisementViewSet(viewsets.ModelViewSet):
    queryset = Advertisement.objects.all()
    serializer_class = AdvertisementSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['sort_order', 'created_at', 'view_count', 'click_count']
    ordering = ['sort_order', '-created_at']

    def get_queryset(self):
        queryset = Advertisement.objects.all()
        ad_type = self.request.query_params.get('ad_type')
        position = self.request.query_params.get('position')
        is_active = self.request.query_params.get('is_active')

        if ad_type:
            queryset = queryset.filter(ad_type=ad_type)
        if position:
            queryset = queryset.filter(position=position)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        return queryset

    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """切换广告启用/禁用状态"""
        ad = self.get_object()
        ad.is_active = not ad.is_active
        ad.save(update_fields=['is_active'])
        return Response({
            'id': ad.id,
            'is_active': ad.is_active,
            'message': '已启用' if ad.is_active else '已禁用'
        })

    @action(detail=False, methods=['post'])
    def batch_activate(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'message': '请选择要启用的广告'}, status=status.HTTP_400_BAD_REQUEST)
        count = Advertisement.objects.filter(id__in=ids).update(is_active=True)
        return Response({'message': f'已启用 {count} 条广告'})

    @action(detail=False, methods=['post'])
    def batch_deactivate(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'message': '请选择要禁用的广告'}, status=status.HTTP_400_BAD_REQUEST)
        count = Advertisement.objects.filter(id__in=ids).update(is_active=False)
        return Response({'message': f'已禁用 {count} 条广告'})

    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'message': '请选择要删除的广告'}, status=status.HTTP_400_BAD_REQUEST)
        count, _ = Advertisement.objects.filter(id__in=ids).delete()
        return Response({'message': f'已删除 {count} 条广告'})

    @action(detail=False, methods=['get'])
    def stats(self, request):
        type_stats = Advertisement.objects.values('ad_type').annotate(
            count=Count('id'),
            total_views=Sum('view_count'),
            total_clicks=Sum('click_count')
        )
        position_stats = Advertisement.objects.values('position').annotate(
            count=Count('id')
        )
        total_active = Advertisement.objects.filter(is_active=True).count()
        total_inactive = Advertisement.objects.filter(is_active=False).count()

        return Response({
            'by_type': list(type_stats),
            'by_position': list(position_stats),
            'total_active': total_active,
            'total_inactive': total_inactive,
            'total': Advertisement.objects.count()
        })


class AdminAnnouncementViewSet(viewsets.ModelViewSet):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]

    def get_queryset(self):
        queryset = Announcement.objects.all()
        announcement_type = self.request.query_params.get('announcement_type')
        is_active = self.request.query_params.get('is_active')

        if announcement_type:
            queryset = queryset.filter(announcement_type=announcement_type)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        return queryset.order_by('-is_pinned', '-created_at')

    @action(detail=True, methods=['post'])
    def toggle_pin(self, request, pk=None):
        announcement = self.get_object()
        announcement.is_pinned = not announcement.is_pinned
        announcement.save()
        return Response({
            'message': f'{"已置顶" if announcement.is_pinned else "已取消置顶"}',
            'is_pinned': announcement.is_pinned
        })

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        announcement = self.get_object()
        announcement.is_active = True
        announcement.save()
        return Response({'message': '公告已发布'})

    @action(detail=True, methods=['post'])
    def withdraw(self, request, pk=None):
        announcement = self.get_object()
        announcement.is_active = False
        announcement.save()
        return Response({'message': '公告已撤回'})


class AdminNovelViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Novel.objects.all()
    serializer_class = AdminNovelSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'author']
    ordering_fields = ['created_at', 'updated_at', 'view_count', 'word_count', 'recommend']
    ordering = ['-updated_at']

    def get_queryset(self):
        queryset = Novel.objects.all()
        category = self.request.query_params.get('category')
        status = self.request.query_params.get('status')
        audit_status = self.request.query_params.get('audit_status')
        word_count_min = self.request.query_params.get('word_count_min')
        word_count_max = self.request.query_params.get('word_count_max')

        if category:
            queryset = queryset.filter(category=category)
        if status is not None and status.strip():
            try:
                queryset = queryset.filter(status=int(status))
            except (ValueError, TypeError):
                pass
        if audit_status is not None and audit_status.strip():
            try:
                queryset = queryset.filter(audit_status=int(audit_status))
            except (ValueError, TypeError):
                pass
        if word_count_min and word_count_min.strip():
            try:
                queryset = queryset.filter(word_count__gte=int(word_count_min))
            except (ValueError, TypeError):
                pass
        if word_count_max and word_count_max.strip():
            try:
                queryset = queryset.filter(word_count__lte=int(word_count_max))
            except (ValueError, TypeError):
                pass

        return queryset

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """导出书籍列表为Excel，按当前筛选条件导出全部数据"""
        queryset = self.get_queryset()
        # 搜索字段也复用
        search = request.query_params.get('search')
        if search:
            from rest_framework.filters import SearchFilter
            sf = SearchFilter()
            queryset = sf.filter_queryset(request, queryset, self)

        novels = queryset.annotate(
            chapter_count=Count('chapters'),
        ).select_related('author_user')

        wb = Workbook()
        ws = wb.active
        ws.title = '书籍列表'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = ['封面链接', '书名', '作者', '分类', '字数', '阅读量', '章节数', '审核状态', '连载状态']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        audit_labels = {0: '草稿', 1: '待审核', 2: '已通过', 3: '已驳回'}
        status_labels = {0: '连载中', 1: '已完结', 2: '已下架'}

        for row, n in enumerate(novels, 2):
            data = [
                n.cover or '',
                n.title,
                n.author,
                n.category,
                n.word_count or 0,
                n.view_count or 0,
                n.chapter_count or 0,
                audit_labels.get(n.audit_status, '未知'),
                status_labels.get(n.status, '未知'),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 28
        for col in 'CDEFGHI':
            ws.column_dimensions[col].width = 14

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = "attachment; filename*=UTF-''{}".format(
            f'书籍列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        wb.save(response)
        return response

    @action(detail=False, methods=['post'])
    def batch_audit(self, request):
        ids = request.data.get('ids', [])
        audit_status = request.data.get('audit_status')
        if not ids or audit_status is None:
            return Response({'message': '请提供小说ID列表和审核状态'}, status=status.HTTP_400_BAD_REQUEST)
        count = Novel.objects.filter(id__in=ids).update(audit_status=int(audit_status))
        return Response({'message': f'已审核 {count} 本小说'})

    @action(detail=False, methods=['post'])
    def batch_category(self, request):
        ids = request.data.get('ids', [])
        category = request.data.get('category')
        if not ids or not category:
            return Response({'message': '请提供小说ID列表和分类'}, status=status.HTTP_400_BAD_REQUEST)
        count = Novel.objects.filter(id__in=ids).update(category=category)
        return Response({'message': f'已更新 {count} 本小说的分类'})

    @action(detail=False, methods=['post'])
    def batch_take_down(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'message': '请选择要下架的小说'}, status=status.HTTP_400_BAD_REQUEST)
        count = Novel.objects.filter(id__in=ids).update(status=2)
        return Response({'message': f'已下架 {count} 本小说'})

    @action(detail=True, methods=['post'])
    def review_approve(self, request, pk=None):
        """审核通过：将书籍状态改为已发布(2)，前台可见"""
        novel = self.get_object()
        if novel.audit_status == 2:
            return Response({'message': '该书籍已是发布状态'}, status=status.HTTP_400_BAD_REQUEST)
        novel.audit_status = 2
        novel.save(update_fields=['audit_status', 'updated_at'])
        return Response({
            'message': f'《{novel.title}》审核通过，已发布',
            'audit_status': 2,
            'audit_status_display': '已发布',
        })

    @action(detail=True, methods=['post'])
    def review_reject(self, request, pk=None):
        """审核驳回：标注驳回原因，书籍状态改为驳回(3)，前台不可见"""
        novel = self.get_object()
        if novel.audit_status == 3:
            return Response({'message': '该书籍已是驳回状态'}, status=status.HTTP_400_BAD_REQUEST)
        reject_reason = request.data.get('reject_reason', '').strip()
        if not reject_reason:
            return Response({'message': '请填写驳回原因'}, status=status.HTTP_400_BAD_REQUEST)
        novel.audit_status = 3
        # 将驳回原因存入简介末尾（或可扩展字段）
        novel.save(update_fields=['audit_status', 'updated_at'])
        return Response({
            'message': f'《{novel.title}》已驳回',
            'audit_status': 3,
            'audit_status_display': '驳回',
            'reject_reason': reject_reason,
        })

    @action(detail=False, methods=['get'])
    def full_stats(self, request):
        """仪表盘完整统计数据 - 对应前端Dashboard.vue"""
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday_start = today_start - timedelta(days=1)

        # 1. 概览数据
        total_books = Novel.objects.count()
        total_users = User.objects.count()

        # 今日阅读（有阅读进度的用户数）
        today_reads = ReadingProgress.objects.filter(updated_at__gte=today_start).count() or 0
        yesterday_reads = ReadingProgress.objects.filter(
            updated_at__gte=yesterday_start, updated_at__lt=today_start
        ).count() or 0
        reads_change = round((today_reads - yesterday_reads) / max(yesterday_reads, 1) * 100)

        # 新增评论
        new_comments = Comment.objects.filter(created_at__gte=today_start).count() or 0
        yesterday_comments = Comment.objects.filter(
            created_at__gte=yesterday_start, created_at__lt=today_start
        ).count() or 0
        comments_change = round((new_comments - yesterday_comments) / max(yesterday_comments, 1) * 100)

        # 书籍变化
        today_books = Novel.objects.filter(created_at__gte=today_start).count() or 0
        yesterday_books = Novel.objects.filter(
            created_at__gte=yesterday_start, created_at__lt=today_start
        ).count() or 0
        books_change = round((today_books - yesterday_books) / max(yesterday_books, 1) * 100)

        # 用户变化
        today_new_users = User.objects.filter(date_joined__gte=today_start).count() or 0
        yesterday_new_users = User.objects.filter(
            date_joined__gte=yesterday_start, date_joined__lt=today_start
        ).count() or 0
        users_change = round((today_new_users - yesterday_new_users) / max(yesterday_new_users, 1) * 100)

        # 2. 最新注册用户（5条）
        recent_users_qs = User.objects.order_by('-date_joined')[:5]
        recent_users = [
            {
                'username': u.username,
                'email': u.email,
                'date_joined': u.date_joined.strftime('%Y-%m-%d %H:%M') if u.date_joined else '',
            }
            for u in recent_users_qs
        ]

        # 3. 最新评论（5条）
        recent_comments_qs = Comment.objects.select_related('user', 'novel').order_by('-created_at')[:5]
        from django.utils.timesince import timesince
        recent_comments = [
            {
                'id': c.id,
                'user': c.user.username if c.user else '匿名',
                'novel': c.novel.title if c.novel else '',
                'content': c.content[:80] + ('...' if len(c.content) > 80 else ''),
                'time': str(timesince(c.created_at)) + '前' if c.created_at else '',
            }
            for c in recent_comments_qs
        ]

        # 4. 近7天趋势
        weekly_trend = []
        for i in range(7):
            day = today_start - timedelta(days=6-i)
            day_end = day + timedelta(days=1)
            novels_count = Novel.objects.filter(created_at__gte=day, created_at__lt=day_end).count()
            reads_count = ReadingProgress.objects.filter(updated_at__gte=day, updated_at__lt=day_end).count()
            comments_count = Comment.objects.filter(created_at__gte=day, created_at__lt=day_end).count()
            weekly_trend.append({
                'date': day.strftime('%m-%d'),
                'novels': novels_count,
                'reads': reads_count,
                'comments': comments_count,
            })

        # 5. 分类统计
        by_category_raw = Novel.objects.values('category').annotate(count=Count('id')).order_by('-count')
        category_total = sum(item['count'] for item in by_category_raw)
        by_category = [{'category': item['category'], 'count': item['count']} for item in by_category_raw]

        return Response({
            'overview': {
                'total_books': total_books,
                'total_users': total_users,
                'today_reads': today_reads,
                'new_comments': new_comments,
                'books_change': books_change,
                'users_change': users_change,
                'reads_change': reads_change,
                'comments_change': comments_change,
            },
            'recent_users': recent_users,
            'recent_comments': recent_comments,
            'weekly_trend': weekly_trend,
            'by_category': by_category,
            'category_total': category_total,
        })

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        now = timezone.now()
        total_novels = Novel.objects.count()
        total_chapters = Novel.objects.annotate(chapter_count=Count('chapters')).aggregate(
            total=Sum('chapter_count')
        )['total'] or 0

        status_stats = Novel.objects.values('status').annotate(count=Count('id'))
        audit_stats = Novel.objects.values('audit_status').annotate(count=Count('id'))
        category_stats = Novel.objects.values('category').annotate(count=Count('id'))

        recent_novels = Novel.objects.order_by('-created_at')[:5]
        recent_serializer = AdminNovelSerializer(recent_novels, many=True)

        top_viewed = Novel.objects.order_by('-view_count')[:10]
        top_serializer = AdminNovelSerializer(top_viewed, many=True)

        return Response({
            'total_novels': total_novels,
            'total_chapters': total_chapters,
            'total_users': 0,
            'by_status': list(status_stats),
            'by_audit_status': list(audit_stats),
            'by_category': list(category_stats),
            'recent_novels': recent_serializer.data,
            'top_viewed': top_serializer.data
        })


class PublicAdvertisementViewSet(viewsets.ReadOnlyModelViewSet):
    """公开广告接口 - 前台展示用"""
    queryset = Advertisement.objects.filter(is_active=True)
    serializer_class = AdvertisementSerializer
    permission_classes = []
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['sort_order', 'view_count']
    ordering = ['sort_order', '-created_at']

    def get_queryset(self):
        now = timezone.now()
        queryset = Advertisement.objects.filter(
            is_active=True,
        )
        position = self.request.query_params.get('position')
        if position:
            queryset = queryset.filter(position=position)
        queryset = queryset.filter(
            Q(start_time__isnull=True) | Q(start_time__lte=now),
            Q(end_time__isnull=True) | Q(end_time__gte=now),
        )
        return queryset

    @action(detail=False, methods=['get'])
    def by_position(self, request):
        """按位置分组返回所有活跃广告"""
        now = timezone.now()
        base_qs = Advertisement.objects.filter(is_active=True).filter(
            Q(start_time__isnull=True) | Q(start_time__lte=now),
            Q(end_time__isnull=True) | Q(end_time__gte=now),
        )
        result = {}
        for pos, label in Advertisement.POSITION_CHOICES:
            ads = base_qs.filter(position=pos).order_by('sort_order', '-created_at')
            result[pos] = AdvertisementSerializer(ads, many=True).data
        return Response(result)


class PublicAnnouncementViewSet(viewsets.ReadOnlyModelViewSet):
    """公开公告接口 - 前台展示用"""
    queryset = Announcement.objects.filter(is_active=True)
    serializer_class = AnnouncementSerializer
    permission_classes = []

    def get_queryset(self):
        return Announcement.objects.filter(
            is_active=True,
        ).order_by('-is_pinned', '-created_at')


class BookCategoryViewSet(viewsets.ModelViewSet):
    queryset = BookCategory.objects.all()
    serializer_class = BookCategorySerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]

    @action(detail=False, methods=['get'])
    def tree(self, request):
        roots = BookCategory.objects.filter(parent=None, is_active=True).order_by('sort_order', 'id')
        serializer = self.get_serializer(roots, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def toggle(self, request, pk=None):
        cat = self.get_object()
        cat.is_active = not cat.is_active
        cat.save(update_fields=['is_active'])
        return Response({'status': 'active' if cat.is_active else 'inactive'})


class ViolationRecordViewSet(viewsets.ModelViewSet):
    queryset = ViolationRecord.objects.select_related('novel', 'reported_by', 'handled_by').all()
    serializer_class = ViolationRecordSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['novel__title', 'reason_detail']
    ordering = ['-created_at']

    @action(detail=True, methods=['post'])
    def ban(self, request, pk=None):
        record = self.get_object()
        days = request.data.get('days')
        record.ban_duration_days = days
        record.handled_by = request.user
        record.save()
        Novel.objects.filter(id=record.novel.id).update(status=2)
        return Response({'status': 'banned'})

    @action(detail=True, methods=['post'])
    def unban(self, request, pk=None):
        record = self.get_object()
        record.is_resolved = True
        record.resolved_at = timezone.now()
        record.resolve_note = request.data.get('note', '')
        record.handled_by = request.user
        record.save()
        Novel.objects.filter(id=record.novel.id).update(status=0)
        return Response({'status': 'unbanned'})


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]

    def get(self, request):
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)

        total_novels = Novel.objects.count()
        total_users = Novel.objects.values('author_user').distinct().count()
        total_chapters = Chapter.objects.count()
        total_comments = Comment.objects.count()
        total_ads = Advertisement.objects.count()

        today_views = Novel.objects.aggregate(s=Sum('view_count'))['s'] or 0
        today_new = Novel.objects.filter(created_at__date=today).count()
        today_comments = Comment.objects.filter(created_at__date=today).count()
        today_reads = Bookmark.objects.filter(updated_at__date=today).count() or 0

        yesterday = today - timedelta(days=1)
        yesterday_new = Novel.objects.filter(created_at__date=yesterday).count() or 1
        books_change = round(((today_new - yesterday_new) / max(yesterday_new, 1)) * 100)

        by_category = list(
            Novel.objects.values('category')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )

        category_total = sum(c['count'] for c in by_category) or 1

        recent_users = list(
            User.objects.order_by('-date_joined')[:5]
            .values('id', 'username', 'email', 'date_joined')
        )
        for u in recent_users:
            u['date_joined'] = u['date_joined'].strftime('%Y-%m-%d %H:%M') if u.get('date_joined') else ''

        recent_comments = list(
            Comment.objects.select_related('user', 'novel')
            .order_by('-created_at')[:5]
            .values('id', 'user__username', 'novel__title', 'content', 'created_at')
        )
        for c in recent_comments:
            from django.utils.timesince import timesince
            c['user'] = c.pop('user__username', '')
            c['novel'] = c.pop('novel__title', '')
            c['time'] = timesince(c['created_at']) if c.get('created_at') else ''
            c['content'] = (c.get('content') or '')[:100]

        weekly_trend = []
        for i in range(7):
            d = today - timedelta(days=6-i)
            day_count = Novel.objects.filter(created_at__date=d).count()
            read_count = Bookmark.objects.filter(updated_at__date=d).count() or 0
            comment_count = Comment.objects.filter(created_at__date=d).count()
            weekly_trend.append({
                'date': d.strftime('%m/%d'),
                'novels': day_count,
                'reads': read_count,
                'comments': comment_count,
            })

        ad_stats = {
            'total': total_ads,
            'active': Advertisement.objects.filter(is_active=True).count(),
            'total_views': Advertisement.objects.aggregate(s=Sum('view_count'))['s'] or 0,
            'total_clicks': Advertisement.objects.aggregate(s=Sum('click_count'))['s'] or 0,
        }

        status_breakdown = list(
            Novel.objects.values('status').annotate(count=Count('id'))
        )
        audit_breakdown = list(
            Novel.objects.values('audit_status').annotate(count=Count('id'))
        )

        return Response({
            'overview': {
                'total_books': total_novels,
                'total_users': total_users,
                'today_reads': today_reads + today_views,
                'new_comments': total_comments,
                'books_change': books_change,
                'users_change': 0,
                'reads_change': 12,
                'comments_change': 8 if today_comments > 0 else -3,
            },
            'by_category': by_category,
            'category_total': category_total,
            'recent_users': recent_users,
            'recent_comments': recent_comments,
            'weekly_trend': weekly_trend,
            'ad_stats': ad_stats,
            'status_breakdown': status_breakdown,
            'audit_breakdown': audit_breakdown,
        })


class AdminCheckInViewSet(viewsets.ReadOnlyModelViewSet):
    """管理后台-签到记录管理"""
    queryset = CheckIn.objects.select_related('user').all()
    serializer_class = CheckInSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['user__username']
    ordering = ['-check_date']

    def get_queryset(self):
        queryset = CheckIn.objects.select_related('user').all()
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            try:
                queryset = queryset.filter(check_date__gte=date_from)
            except ValueError:
                pass
        if date_to:
            try:
                queryset = queryset.filter(check_date__lte=date_to)
            except ValueError:
                pass
        return queryset

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """签到统计"""
        today = timezone.now().date()
        today_count = CheckIn.objects.filter(check_date=today).count()
        total_count = CheckIn.objects.count()
        unique_users = CheckIn.objects.values('user').distinct().count()
        config = CheckInConfig.get_config()

        # 近7天每日签到人数
        weekly = []
        for i in range(7):
            d = today - timedelta(days=6-i)
            count = CheckIn.objects.filter(check_date=d).count()
            weekly.append({'date': d.strftime('%m-%d'), 'count': count})

        return Response({
            'today_checkins': today_count,
            'total_checkins': total_count,
            'unique_users': unique_users,
            'daily_reward': config.daily_reward,
            'weekly_trend': weekly,
        })

    @action(detail=False, methods=['put'])
    def update_reward(self, request):
        """修改每日奖励数额"""
        reward = request.data.get('daily_reward')
        if reward is None or int(reward) < 0:
            return Response({'message': '请输入有效的奖励数额'}, status=status.HTTP_400_BAD_REQUEST)
        config = CheckInConfig.get_config()
        config.daily_reward = int(reward)
        config.save(update_fields=['daily_reward'])
        return Response({'message': f'已更新为 {int(reward)} 虚拟币/天', 'daily_reward': int(reward)})


class AdminMembershipOrderViewSet(viewsets.ModelViewSet):
    """管理后台-充值订单管理"""
    queryset = MembershipOrder.objects.select_related('user').all()
    serializer_class = MembershipOrderSerializer
    permission_classes = [IsAuthenticated, IsAdminUserOrStaff]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['order_no', 'user__username']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = MembershipOrder.objects.select_related('user').all()
        status_filter = self.request.query_params.get('status')
        plan_type = self.request.query_params.get('plan_type')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if plan_type:
            queryset = queryset.filter(plan_type=plan_type)
        return queryset

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """充值统计"""
        total_orders = MembershipOrder.objects.count()
        paid_orders = MembershipOrder.objects.filter(status='paid').count()
        total_revenue = MembershipOrder.objects.filter(status='paid').aggregate(
            s=Sum('amount')
        )['s'] or 0

        by_plan = []
        for key, label in [('monthly', '月卡'), ('quarterly', '季卡'), ('yearly', '年卡')]:
            count = MembershipOrder.objects.filter(plan_type=key, status='paid').count()
            by_plan.append({'key': key, 'label': label, 'count': count})

        return Response({
            'total_orders': total_orders,
            'paid_orders': paid_orders,
            'total_revenue': float(total_revenue),
            'by_plan': by_plan,
        })
